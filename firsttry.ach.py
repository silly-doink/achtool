import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def browse_file():
    filename = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if filename:
        # Read TXT file
        with open(filename, 'r') as txt_file:
            data = txt_file.read()
        # Convert and highlight data
        xlsx_file_path = convert_and_highlight(filename, data)
        if xlsx_file_path:
            status_label.config(text=f"XLSX file created: {xlsx_file_path}")
        else:
            status_label.config(text="Failed to create XLSX file.")

def highlight_cell(value, row_num, col_num, total_rows):
    # Define highlighting rules for the first row (header)
    if row_num == 1 and (2 <= col_num <= 10 or 21 <= col_num <= 24 or 31 <= col_num <= 35 or 56 <= col_num <= 58):
        return value, 'FF0000'  # Red color

    # Define highlighting rules for subsequent rows
    elif row_num > 1 and row_num < total_rows:
        if (2 <= col_num <= 10 or 21 <= col_num <= 24 or 28 <= col_num <= 37 
            or 39 <= col_num <= 40 or 44 <= col_num <= 47 or 53 <= col_num <= 64
            or 87 <= col_num <= 89 or 105 <= col_num <= 134 or 165 <= col_num <= 174
            or 194 <= col_num <= 202 or 215 <= col_num <= 229 or 252 <= col_num <= 253):
            return value, '00FF00'  # Green color
        
    # Define highlighting rules for the last row
    elif row_num == total_rows and (2 <= col_num <= 10 or 21 <= col_num <= 24 or 39 <= col_num <= 46 
                                    or 61 <= col_num <= 68 or 83 <= col_num <= 90 or 105 <= col_num <= 112):
        return value, 'FFFF00'  # Yellow color
    
    # Default color (white) if no rule matches
    return value, 'FFFFFF' 


def convert_and_highlight(txt_file_path, data):
    try:
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Total number of rows in the data
        total_rows = len(data.splitlines())

        # Apply highlighting rules to each character in the data
        for row_num, row in enumerate(data.splitlines(), start=1):
            for col_num, char in enumerate(row, start=1):
                highlighted_char, color = highlight_cell(char, row_num, col_num, total_rows)
                cell = ws.cell(row=row_num, column=col_num, value=highlighted_char)
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill

        # Save the workbook to an XLSX file in the user's Documents directory
        documents_path = os.path.join(os.path.expanduser("~"), "Documents")
        xlsx_file_path = os.path.join(documents_path, os.path.basename(txt_file_path).replace('.txt', '_highlighted.xlsx'))
        wb.save(xlsx_file_path)

        # Print the save path for debugging
        print(f"File saved to: {xlsx_file_path}")

        return xlsx_file_path
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# Create the main application window
root = tk.Tk()
root.title("TXT to XLSX Highlighter")

# Create a button to browse for a TXT file
browse_button = tk.Button(root, text="Browse for TXT File", command=browse_file)
browse_button.pack(pady=10)

# Create a label to display status messages
status_label = tk.Label(root, text="")
status_label.pack()

# Run the Tkinter event loop
root.mainloop()

